# =================================================================================================================
# Unified API Framework with CLI-Driven Excel Data Generation
# Created by John Panicker
# Usage for test execution : python -m pytest api_framework_tests_v2.py -q --tb=short --alluredir=../allure-results 
# Usage for test data generation: python -m pytest api_framework_tests_v2.py --generate-data
# ==================================================================================================================

# Import relevant classes and functions.

import json
import os
import datetime
import random
import requests
import pytest
import allure

from dataclasses import dataclass
from typing import Optional, Dict
from requests.auth import HTTPBasicAuth
from openpyxl import load_workbook, Workbook
from data_sampler import generate_samples


# ==============================================================================================================================
# Global Configuration
# Reads existing test data from ../test-data/sscData.xlsx file for test execution.
# If generate column in the sscData.xlsx is populated as Y and if we are using --generate-data in the CLI option as above, 
# the script will analyse the number of entries and requests info returned from the API requests and generate the test data file 
# from those request response as /test-data/sscData_generated.xlsx
# ================================================================================================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

TEST_DATA_FILE = os.path.join(
    BASE_DIR, "../test-data/sscData.xlsx"
)

GENERATED_FILE = os.path.join(
    BASE_DIR, "../test-data/sscData_generated.xlsx"
)

MAX_SAMPLE_ELEMENTS = 3

# ============================================================
# OAuth Configuration (Client Credentials) with tokens
# ============================================================

TOKEN_URL = "https://loyalty-qa.app.fielo.com/dx/uas/oauth/token"
CLIENT_ID = "xqmZxEEi07svmnTk"
CLIENT_SECRET = "tfBACoKc4VTHQLCeAE27L3CKySGCinl6"

_cached_oauth_token = None


def get_oauth_token() -> str:
    """Fetch and cache OAuth token using client credentials."""
    global _cached_oauth_token

    if _cached_oauth_token:
        return _cached_oauth_token

    payload = {
        "grant_type": "client_credentials",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
    }

    response = requests.post(TOKEN_URL, data=payload)
    response.raise_for_status()

    token = response.json().get("access_token")

    if not token:
        raise RuntimeError("OAuth token missing")

    _cached_oauth_token = token
    return token

# ============================================================
# Environment Resolution
# ============================================================

ENVIRONMENT_URLS = {
    "local": "http://localhost",
    "dev": "https://awddev.trialclient1.awdcloud.co.uk",
    "lunate": "https://awddev.lunate.ae-bpchorus.com",
    "prod": "https://awdprod.trialclient1.awdcloud.co.uk"
}


# ============================================================
# Allure Environment
# ============================================================

_env_file_created = False


def create_allure_environment_file(env, url, user):

    global _env_file_created

    if _env_file_created:
        return

    _env_file_created = True

    results_dir = os.path.join(os.getcwd(), "allure-results")
    os.makedirs(results_dir, exist_ok=True)

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    content = (
        f"Environment={env}\n"
        f"BaseURL={url}\n"
        f"User={user or 'N/A'}\n"
        f"ExecutionTime={now}\n"
    )

    with open(
        os.path.join(results_dir, "environment.properties"),
        "w",
        encoding="utf-8"
    ) as f:
        f.write(content)


# ============================================================
# Excel Utilities
# ============================================================

# Change index if needed as per the data is coming from the Excel spreadsheet.Sheet number starts at 0.

SHEET_INDEX = 8

# =======================================================================
#  Excel cells sometimes come in as None, "none", "nan" if it is blank.
#  To be safe, you should explicitly check that it’s non-empty
#  and not the string "None".
# =======================================================================

def is_valid_check(value) -> bool:
    return value not in (None, "") and str(value).strip().lower() not in ("none", "nan")


def get_test_data(test_case_id: str) -> Optional[dict]:

    wb = load_workbook(TEST_DATA_FILE, data_only=True)
    sheet = wb.worksheets[SHEET_INDEX]

    headers = [c.value for c in sheet[1]]

    for row in sheet.iter_rows(min_row=2, values_only=True):

        record = dict(zip(headers, row))

        if record.get("TestCase") == test_case_id:
            return record

    return None

# If the test cases needs to be executed enter Y in the Execute column in the test data spreadsheet.

def collect_test_cases() -> list:

    wb = load_workbook(TEST_DATA_FILE, data_only=True)
    sheet = wb.worksheets[SHEET_INDEX]

    headers = [c.value for c in sheet[1]]

    tc = headers.index("TestCase")
    ex = headers.index("Execute")

    result = []

    for row in sheet.iter_rows(min_row=2, values_only=True):

        if (
            is_valid_check(row[tc])
            and str(row[ex]).upper() == "Y"
        ):
            result.append(row[tc])

    return result


# ============================================================
# Excel Generation Helpers
# ============================================================

def clone_excel():

    if os.path.exists(GENERATED_FILE):
        return

    src = load_workbook(TEST_DATA_FILE)
    dst = Workbook()

    src_sheet = src.worksheets[SHEET_INDEX]
    dst_sheet = dst.active
    dst_sheet.title = src_sheet.title

    for row in src_sheet.iter_rows(values_only=True):
        dst_sheet.append(row)

    dst.save(GENERATED_FILE)


def update_generated_file(
    test_case,
    count,
    elements
):

    wb = load_workbook(GENERATED_FILE)
    sheet = wb.active

    headers = [c.value for c in sheet[1]]

    tc_idx = headers.index("TestCase") + 1
    count_idx = headers.index("check_count") + 1

    elem_cols = {
        h: headers.index(h) + 1
        for h in headers
        if h.startswith("check_element")
    }

    for row in sheet.iter_rows(min_row=2):

        if row[tc_idx - 1].value == test_case:

            row[count_idx - 1].value = count

            for i, val in enumerate(elements):

                key = f"check_element{i+1}"

                if key in elem_cols:
                    cell_index = elem_cols[key] - 1

                    # If value is a list, spread across max 3 columns
                    if isinstance(val, list):
                        for offset, item in enumerate(val[:3]):
                            col = cell_index + offset
                            # Do not write past check_element3
                            if col <= elem_cols["check_element3"] - 1:
                                row[col].value = item
                    else:
                        row[cell_index].value = val
            break

    wb.save(GENERATED_FILE)


# ================================================================================================================================================================
# URL Builder. If the excel data has base_url starting with http://localhost and env variable has a value in Excel, 
# then local host will get replaced by what is mentioned in the env variable. For e.g. base_url is http://localhost/awdServer/awd/services/v1/user/businessareas
# and environemnt variable is dev, eventually it will become https://awddev.trialclient1.awdcloud.co.uk/awdServer/awd/services/v1/user/businessareas
# ================================================================================================================================================================

def build_final_url(base_url, env, test_data):

    env_base = ENVIRONMENT_URLS.get(env, ENVIRONMENT_URLS["local"])

    if base_url.startswith("http://localhost"):
        base_url = base_url.replace(
            "http://localhost", env_base
        )

    for k, v in test_data.items():

        if is_valid_check(v):
            base_url = base_url.replace(
                f"{{{k}}}", str(v)
            )

    return (
        base_url
        .replace("://", "__TMP__")
        .replace("//", "/")
        .replace("__TMP__", "://")
    )


# ======================================================================================================
# Unified Authentication Abstraction - combining the HTTPBasicAuth and Explicit Bearer token from Excel
# ======================================================================================================

@dataclass
class AuthContext:
    headers: Dict[str, str]
    auth: Optional[HTTPBasicAuth]


def resolve_auth_context(test_data):

    """
    Auth resolution priority:

    1. HTTP Basic Auth (username + password)
    2. Explicit Bearer token from Excel
    3. No authentication (if all are blank)
    4. OAuth Client Credentials (fallback)
    """

    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    token = test_data.get("token")
    user = test_data.get("username")
    pwd = test_data.get("password")

    # --------------------------------------------------
    # 1. HTTP Basic Authentication
    # --------------------------------------------------
    if is_valid_check(user) and is_valid_check(pwd):

        return AuthContext(
            headers=headers,
            auth=HTTPBasicAuth(user, pwd)
        )
    # --------------------------------------------------
    # 2. Explicit Bearer Token
    # --------------------------------------------------
    if is_valid_check(token):
        headers["Authorization"] = f"Bearer {token}"
        return AuthContext(headers, None)
    
    # --------------------------------------------------
    # 3. No Authentication
    # --------------------------------------------------
    if not is_valid_check(user) and not is_valid_check(token):
        return AuthContext(headers,None)
    
    # --------------------------------------------------
    # 4. OAuth Client Credentials (Fallback)
    # --------------------------------------------------
    oauth = get_oauth_token()
    headers["Authorization"] = f"Bearer {oauth}"

    return AuthContext(headers, None)

# ============================================================
# Test Collection
# ============================================================

test_cases = collect_test_cases()


# ============================================================
# Main Test
# ============================================================

@pytest.mark.parametrize("test_case_id", test_cases)
@allure.feature("Unified API")
def test_api_validation(
    test_case_id,
    generate_data_enabled
):

    test_data = get_test_data(test_case_id)

    assert test_data, f"No data for {test_case_id}"

    method = str(
        test_data.get("method", "GET")
    ).upper()

    base_url = build_final_url(
        test_data.get("base_url"),
        str(test_data.get("environment", "dev")).lower(),
        test_data
    )

    auth = resolve_auth_context(test_data)

    payload = test_data.get("payload")
    data = None

    if is_valid_check(payload):
        data = json.loads(payload)
        allure.attach(payload, "Payload", allure.attachment_type.TEXT)

    allure.attach(base_url, "Resolved API Endpoint", allure.attachment_type.TEXT)

    create_allure_environment_file(
        test_data.get("environment"),
        base_url,
        test_data.get("username")
    )

    with allure.step(f"Sending {method} request"):
        response = requests.request(
            method=method,
            url=base_url,
            headers=auth.headers,
            auth=auth.auth,
            json=data
        )

    allure.attach(str(response.status_code), "HTTP Status Code", allure.attachment_type.TEXT)
    allure.attach(response.text, "Raw Response", allure.attachment_type.TEXT)

    # --------------------------------------------------
    # Fail test on HTTP error responses (4xx / 5xx)
    # --------------------------------------------------
    if response.status_code >= 400:
        error_message = None

        try:
            error_body = response.json()
            error_message = error_body.get("errors") or error_body.get("message")
        except ValueError:
            error_body = response.text

        pytest.fail(
            f"API request failed.\n"
            "Status Code: {response.status_code}\n"
            f"Response: {error_body}"
        )

    info = None

    try:
        info = response.json()
    except Exception:
        pass


    # ========================================================
    # Test Data Automatic Generation Logic
    # ========================================================

    generate_flag = str(
        test_data.get("generate", "")
    ).upper()

    gen_count = 0
    samples = []

    if (
        generate_data_enabled
        and generate_flag == "Y"
        and isinstance(info, dict)
    ):

        entries = next(
            (
                v for v in info.values()
                if isinstance(v, list)
            ),
            []
        )

        if entries:

            clone_excel()

            samples = generate_samples(
                entries,
                max_samples=MAX_SAMPLE_ELEMENTS
            )

            gen_count = len(samples)

    formatted = []

    for row in samples:
        formatted.append([str(v) for v in row])

    if formatted:
        update_generated_file(
            test_case_id,
            gen_count,
            formatted
        )


    # ========================================================
    # Validation Logic
    # ========================================================

    check_count = test_data.get("check_count")

    if is_valid_check(check_count) and info:

        entries = next(
            (
                v for v in info.values()
                if isinstance(v, list)
            ),
            []
        )

        assert len(entries) >= int(check_count)
        allure.attach(str(len(entries)), "Total Entries Returned", allure.attachment_type.TEXT)

    for key, raw_value in test_data.items():
        if key.startswith("check_element"):
            value = str(raw_value).strip('"')
            if is_valid_check(value) and info:
                with allure.step(f"Checking presence of '{value}' in response"):
                    assert value in json.dumps(info)

    allure.dynamic.title(f"{test_case_id} - {method} {base_url}")
    allure.dynamic.description("Unified authentication API validation")