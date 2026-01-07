import json
import requests
from requests.auth import HTTPBasicAuth
import os
from openpyxl import load_workbook
import pytest
import allure
import datetime

# ==============================================================
#  Environment mapping for dynamic URL resolution
#  (New: maps environment keywords from Excel to actual base URLs)
# ==============================================================
ENVIRONMENT_URLS = {
    "local": "http://localhost",
    "dev": "https://awddev.trialclient1.awdcloud.co.uk",
    "lunate": "https://awddev.lunate.ae-bpchorus.com",
    "prod": "https://awdprod.trialclient1.awdcloud.co.uk"
}

# ==============================================================
#  Create Allure Environment tab information
#  (New: generates allure-results/environment.properties automatically)
# ==============================================================
_env_file_created = False

def create_allure_environment_file(env_name, base_url, username):
    """
    Creates or updates the Allure environment.properties file.
    This file populates the 'Environment' tab in the Allure report dashboard.
    """
    global _env_file_created
    if _env_file_created:
        return
    _env_file_created = True

    results_dir = os.path.join(os.getcwd(), "allure-results")
    os.makedirs(results_dir, exist_ok=True)

    env_file_path = os.path.join(results_dir, "environment.properties")
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    content = (
        f"Test Environment={env_name}\n"
        f"Base URL={base_url}\n"
        f"Username={username or 'N/A'}\n"
        f"Execution Time={now}\n"
        f"Executed By={os.getenv('USERNAME') or os.getenv('USER') or 'Unknown'}\n"
    )

    with open(env_file_path, "w", encoding="utf-8") as f:
        f.write(content)

# ==============================================================
#  Fetching data from the Excel spreadsheet.
# ==============================================================
def get_test_data(test_case_id):
    """Fetch test case data from Excel."""
    file_path = os.path.join(os.getcwd(), "../", "test-data", "sscData.xlsx")
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.worksheets[3]

    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    data = [
        {headers[i]: cell for i, cell in enumerate(row)}
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]

    for row in data:
        if row.get("TestCase") == test_case_id:
            return row
    return None

# ==============================================================
#  Excel cells sometimes come in as None, "none", "nan" if it is blank.
#  To be safe, you should explicitly check that it’s non-empty
#  and not the string "None".
# ==============================================================
def is_valid_check(value):
    return value not in (None, "") and str(value).strip().lower() not in ["none", "nan"]

# ==============================================================
#  Helper function to dynamically replace placeholders in URL.
#  (Replaces {instanceId}, {businessAreaId}, {typeId}, etc.)
# ==============================================================
def build_final_url(base_url, environment, test_data):
    """
    Replaces 'http://localhost' with environment base URL and
    dynamically substitutes variables like {instanceId}, {businessAreaId}, {typeId}, etc.
    """
    env_base = ENVIRONMENT_URLS.get(environment.lower(), ENVIRONMENT_URLS["local"])

    # Replace localhost with actual environment base
    if base_url.startswith("http://localhost"):
        base_url = base_url.replace("http://localhost", env_base)

    # Replace any placeholders in the URL with corresponding Excel values
    for key, value in test_data.items():
        if is_valid_check(value) and f"{{{key}}}" in base_url:
            base_url = base_url.replace(f"{{{key}}}", str(value))

    # Sanitize accidental double slashes (except after 'https://')
    base_url = base_url.replace("://", "TEMP_PLACEHOLDER").replace("//", "/").replace("TEMP_PLACEHOLDER", "://")

    return base_url

# ==============================================================
#  Dynamically collect test cases from Excel spreadsheet
# ==============================================================
test_cases = [f"TestCase{i}" for i in range(11, 12)]

# ==============================================================
#  Main API validation test
# ==============================================================
@pytest.mark.parametrize("test_case_id", test_cases)
@allure.feature("Chorus SS&C API Regression")
@allure.severity(allure.severity_level.CRITICAL)
def test_api_validation(test_case_id):
    """Main API validation test."""
    test_data = get_test_data(test_case_id)
    assert test_data is not None, f"❌ Test data not found for {test_case_id}"

    # ==============================================================
    #  Gathering data from the input spreadsheet.
    # ==============================================================
    method = str(test_data.get("method", "GET")).upper()
    base_url_from_excel = test_data.get("base_url")
    username = test_data.get("username")
    password = test_data.get("password")
    payload = test_data.get("payload")
    root_element = test_data.get("root_element")
    check_count = test_data.get("check_count")
    check_element1 = str(test_data.get("check_element1")).strip('"')
    check_element2 = str(test_data.get("check_element2")).strip('"')
    environment = str(test_data.get("environment", "dev")).strip().lower()

    # ==============================================================
    #  Dynamically construct the final URL.
    #  (New: replaces localhost + substitutes placeholders from Excel)
    # ==============================================================
    base_url = build_final_url(base_url_from_excel, environment, test_data)

    # ==============================================================
    #  Attach environment and resolved URL to the Allure report.
    #  (New: increases transparency in the test run report)
    # ==============================================================
    allure.attach(environment, name="Environment", attachment_type=allure.attachment_type.TEXT)
    allure.attach(base_url, name="Resolved API Endpoint", attachment_type=allure.attachment_type.TEXT)

    # ==============================================================
    #  Create Allure Environment tab metadata file.
    #  (New: appears as a separate “Environment” tab in Allure dashboard)
    # ==============================================================
    create_allure_environment_file(environment, base_url, username)

    # ==============================================================
    #  Making sure that the request sent is of type JSON.
    # ==============================================================
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json"
    }

    # Convert payload if present
    data = None
    if payload:
        try:
            data = json.loads(payload)
            allure.attach(payload, name="Payload", attachment_type=allure.attachment_type.TEXT)
        except Exception as e:
            pytest.fail(f"Invalid JSON payload in Excel for {test_case_id}: {e}")

    # ==============================================================
    #  Sending the API request to the system.
    #  If username and password are not provided in the Excel spreadsheet,
    #  None is sent in the request.
    # ==============================================================
    with allure.step(f"Sending {method} request to {base_url}"):
        try:
            response = requests.request(
                method=method,
                url=base_url,
                auth=HTTPBasicAuth(username, password) if username and password else None,
                headers=headers,
                json=data  # works for POST/PUT/PATCH automatically
            )
        except Exception as e:
            pytest.fail(f"Request failed: {e}")

    # ==============================================================
    #  Attaching the HTTP Status Code and Raw Response into the Allure report.
    # ==============================================================
    allure.attach(str(response.status_code), "HTTP Status Code", allure.attachment_type.TEXT)
    allure.attach(response.text, "Raw Response", allure.attachment_type.TEXT)

    # ==============================================================
    #  GET, POST, PUT methods return responses from system.
    #  Noticed that DELETE methods are not returning responses.
    #  Hence we are skipping the JSON validation for some methods which
    #  are not returning the response.
    # ==============================================================
    info = None
    if method in ["GET", "POST", "PUT"]:
        with allure.step("Validating JSON response format"):
            try:
                info = response.json()
            except ValueError:
                pytest.fail(f"Response is not valid JSON for {method} request!")
    else:
        with allure.step(f"Skipping JSON validation for {method} request"):
            allure.attach(response.text, name="Raw Response", attachment_type=allure.attachment_type.TEXT)

    # ==============================================================
    #  Optional: Validate root element and count if relevant.
    #  This check will only happen if user inputs data for root_element in the spreadsheet.
    # ==============================================================
    if is_valid_check(root_element) and info is not None:
        try:
            count = len(info[root_element])
        except Exception:
            count = len(info)
        allure.attach(str(count), "Total Entries Returned", allure.attachment_type.TEXT)
        if check_count:
            with allure.step(f"Verifying count >= {check_count}"):
                assert count >= int(check_count), f"Entries ({count}) < expected ({check_count})"

    # ==============================================================
    #  Generic content checks.
    #  These checks will only happen if user inputs data for
    #  check_element1 or check_element2 in the spreadsheet.
    # ==============================================================
    if is_valid_check(check_element1) and info is not None:
        with allure.step(f"Checking presence of '{check_element1}' in response"):
            assert check_element1 in json.dumps(info), f"`{check_element1}` not found in response"

    if is_valid_check(check_element2) and info is not None:
        with allure.step(f"Checking presence of '{check_element2}' in response"):
            assert check_element2 in json.dumps(info), f"`{check_element2}` not found in response"

    # ==============================================================
    #  Providing the title and description of the Allure report test cases.
    # ==============================================================
    allure.dynamic.title(f"Validate API: {test_case_id}-{method} {base_url}")
    allure.dynamic.description(f"Verifies response structure and data for {base_url}")
