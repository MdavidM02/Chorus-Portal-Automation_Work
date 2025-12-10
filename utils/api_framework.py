import json
import requests
from requests.auth import HTTPBasicAuth
import os
from openpyxl import load_workbook

def get_test_data(test_case_id):
    # Build the absolute file path
    file_path = os.path.join(os.getcwd(), "../","test-data", "sscData.xlsx")
    # Load workbook
    workbook = load_workbook(file_path, data_only=True)
    # Select the third sheet with name as API
    sheet = workbook.worksheets[2]
    # Extract header row
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    # Convert sheet to list of dictionaries
    data = [
        {headers[i]: cell for i, cell in enumerate(row)}
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]
    # Find row where TestCase matches test_case_id
    for row in data:
        if row.get("TestCase") == test_case_id:
            return row
    # Return None if not found
    return None

# Fetching data for a specific test case.
test_data = get_test_data("TestCase11")

# Fetching parameters from Excel spreadsheet for a specific test case.
base_url=test_data.get('base_url')
username=test_data.get('username')
password=test_data.get('password')
root_element=test_data.get('root_element')
child_elem1=test_data.get('child_elem1').strip('"')
child_elem2=test_data.get('child_elem2').strip('"')
check_count=test_data.get('check_count')
check_element1=test_data.get('check_element1')
check_element2=test_data.get('check_element2')

headers = {
    "Content-Type": "application/json",
    "Accept": "application/json"  # tells the server we want JSON
}

#parameters = {}
# at the moment, parameters are not used. But we could use those in future.

response = requests.get(base_url,auth=HTTPBasicAuth(username,password),headers=headers) 

try:
    info = response.json()
except ValueError:
    print("Response is not valid JSON!")
    info = None

# Traverse JSON structure
if info and root_element in info:
    for area in info[root_element]:
        type = area.get(child_elem1)
        href = area.get(child_elem2)
        print(f"Type: {type}, Href: {href}")
else:
    print("No 'root_element' found in the API response.")


# Assertions are written in try so that script will continue even if one of the assertions fail.
# Assertion 1 - Checking the total number of entries returned
count = len(info[root_element])
print("Number of entries:", count)
print(check_count)
print(check_element1)
print(check_element2)
print(test_data)

try:
    assert count > int(check_count), "Total number of entries returned should be greater than 9"
except AssertionError as e:
    print(f"Assertion failed: {e}")

if "BALEP" in json.dumps(info):
    print("Found BALEP!")

# Assertion 2 - Checking a specific entry in the response.
try:
    assert check_element1 in json.dumps(info), "No `check_element1` found in response."
except AssertionError as e:
    print(f"Assertion failed: {e}")

# Assertion 3 - Checking a specific entry in the response.
try:
    assert check_element2 in json.dumps(info), "No `check_element2` found in response."
except AssertionError as e:
    print(f"Assertion failed: {e}")