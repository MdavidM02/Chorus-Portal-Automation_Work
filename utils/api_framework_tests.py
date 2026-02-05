# Import relevant classes and functions.

import json
import os
import datetime
import requests
import pytest
import allure

from dataclasses import dataclass
from typing import Optional, Dict
from requests.auth import HTTPBasicAuth
from openpyxl import load_workbook

# OAuth Configuration (Client Credentials) with tokens

TOKEN_URL = "https://loyalty-qa.app.fielo.com/dx/uas/oauth/token"
CLIENT_ID = "xqmZxEEi07svmnTk"
CLIENT_SECRET = "tfBACoKc4VTHQLCeAE27L3CKySGCinl6"
_cached_oauth_token = None

def get_oauth_token() -> str:
    """Fetch and cache OAuth token using client credentials."""
    global _cached_oauth_token

    if _cached_oauth_token:
        return _cached_oauth_token

    payload = {
        "grant_type": "client_credentials",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
    }

    response = requests.post(TOKEN_URL, data=payload)
    response.raise_for_status()

    token = response.json().get("access_token")
    if not token:
        raise RuntimeError("OAuth access token missing in response")
    _cached_oauth_token = token
    return token

# Environment Resolution

ENVIRONMENT_URLS = {
    "local": "http://localhost",
    "dev": "https://awddev.trialclient1.awdcloud.co.uk",
    "lunate": "https://awddev.lunate.ae-bpchorus.com",
    "prod": "https://awdprod.trialclient1.awdcloud.co.uk"
}

# Allure Environment Metadata

_env_file_created = False

def create_allure_environment_file(env_name: str, base_url: str, username: Optional[str]):
    global _env_file_created
    if _env_file_created:
        return

    _env_file_created = True
    results_dir = os.path.join(os.getcwd(), "allure-results")
    os.makedirs(results_dir, exist_ok=True)

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    content = (
        f"Test Environment={env_name}\n"
        f"Base URL={base_url}\n"
        f"Username={username or 'N/A'}\n"
        f"Execution Time={now}\n"
        f"Executed By={os.getenv('USERNAME') or os.getenv('USER') or 'Unknown'}\n"
    )

    with open(os.path.join(results_dir, "environment.properties"), "w", encoding="utf-8") as f:
        f.write(content)


# Excel Utilities

def get_test_data(test_case_id: str) -> Optional[dict]:
    file_path = os.path.join(os.getcwd(), "../", "test-data", "sscData.xlsx")
    workbook = load_workbook(file_path, data_only=True)

    # Change index if needed as per the data is coming from the Excel spreadsheet.Sheet number starts at 0.
    sheet = workbook.worksheets[7]

    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if record.get("TestCase") == test_case_id:
            return record

    return None

# ==============================================================
#  Excel cells sometimes come in as None, "none", "nan" if it is blank.
#  To be safe, you should explicitly check that it’s non-empty
#  and not the string "None".
# ==============================================================

def is_valid_check(value) -> bool:
    return value not in (None, "") and str(value).strip().lower() not in ("none", "nan")

# URL Builder. If the excel data has base_url starting with http://localhost and env variable has a value in Excel, 
# then local host will get replaced by what is mentioned in the env variable. For e.g. base_url is http://localhost/awdServer/awd/services/v1/user/businessareas
# and environemnt variable is dev, eventually it will become https://awddev.trialclient1.awdcloud.co.uk/awdServer/awd/services/v1/user/businessareas

def build_final_url(base_url: str, environment: str, test_data: dict) -> str:
    env_base = ENVIRONMENT_URLS.get(environment.lower(), ENVIRONMENT_URLS["local"])

    if base_url.startswith("http://localhost"):
        base_url = base_url.replace("http://localhost", env_base)

    for key, value in test_data.items():
        if is_valid_check(value):
            base_url = base_url.replace(f"{{{key}}}", str(value))

    return (
        base_url
        .replace("://", "__TMP__")
        .replace("//", "/")
        .replace("__TMP__", "://")
    )


# Unified Authentication Abstraction - combining the HTTPBasicAuth and Explicit Bearer token from Excel

@dataclass
class AuthContext:
    headers: Dict[str, str]
    auth: Optional[HTTPBasicAuth]

def resolve_auth_context(test_data: dict) -> AuthContext:
    """
    Auth resolution priority:

    1. HTTP Basic Auth (username + password)
    2. Explicit Bearer token from Excel
    3. No authentication (if all are blank)
    4. OAuth Client Credentials (fallback)
    """

    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json"
    }

    token = test_data.get("token")
    username = test_data.get("username")
    password = test_data.get("password")

    has_token = is_valid_check(token)
    has_user_pass = is_valid_check(username) and is_valid_check(password)

    # --------------------------------------------------
    # 1. HTTP Basic Authentication
    # --------------------------------------------------
    if has_user_pass:
        return AuthContext(
            headers=headers,
            auth=HTTPBasicAuth(username, password)
        )

    # --------------------------------------------------
    # 2. Explicit Bearer Token
    # --------------------------------------------------
    if has_token:
        headers["Authorization"] = f"Bearer {token}"
        return AuthContext(
            headers=headers,
            auth=None
        )

    # --------------------------------------------------
    # 3. No Authentication
    # --------------------------------------------------
    if not has_user_pass and not has_token:
        return AuthContext(
            headers=headers,
            auth=None
        )

    # --------------------------------------------------
    # 4. OAuth Client Credentials (Fallback)
    # --------------------------------------------------
    oauth_token = get_oauth_token()
    headers["Authorization"] = f"Bearer {oauth_token}"

    return AuthContext(
        headers=headers,
        auth=None
    )


#  Excel-Driven Enable/Disable Flag for test case execution. If Execute variable in Excel is Y, the test gets executed. else it will not.

def collect_test_cases() -> list[str]:
    file_path = os.path.join(os.getcwd(), "../", "test-data", "sscData.xlsx")
    workbook = load_workbook(file_path, data_only=True)
    
    # Change index if needed as per the data is coming from the Excel spreadsheet.Sheet number starts at 0.
    sheet = workbook.worksheets[7]

    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    tc_idx = headers.index("TestCase")
    exec_idx = headers.index("Execute")

    test_cases = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        test_case_id = row[tc_idx]
        execute_flag = row[exec_idx]

        if is_valid_check(test_case_id) and str(execute_flag).strip().upper() == "Y":
            test_cases.append(test_case_id)

    return test_cases

test_cases = collect_test_cases()

# Unified API Test and updating the results to Allure reports.

@pytest.mark.parametrize("test_case_id", test_cases)
@allure.feature("Unified API Regression")
@allure.severity(allure.severity_level.CRITICAL)
def test_api_validation(test_case_id):
    test_data = get_test_data(test_case_id)
    assert test_data, f"Test data not found for {test_case_id}"

    method = str(test_data.get("method", "GET")).upper()
    base_url = build_final_url(
        test_data.get("base_url"),
        str(test_data.get("environment", "dev")).lower(),
        test_data
    )

    auth_context = resolve_auth_context(test_data)

    payload = test_data.get("payload")
    data = None
    if is_valid_check(payload):
        try:
            data = json.loads(payload)
            allure.attach(payload, "Payload", allure.attachment_type.TEXT)
        except Exception as e:
            pytest.fail(f"Invalid JSON payload: {e}")

    allure.attach(base_url, "Resolved API Endpoint", allure.attachment_type.TEXT)
    
    create_allure_environment_file(
        test_data.get("environment"),
        base_url,
        test_data.get("username")
    )

    with allure.step(f"Sending {method} request"):
        response = requests.request(
            method=method,
            url=base_url,
            headers=auth_context.headers,
            auth=auth_context.auth,
            json=data
        )

    allure.attach(str(response.status_code), "HTTP Status Code", allure.attachment_type.TEXT)
    allure.attach(response.text, "Raw Response", allure.attachment_type.TEXT)

    # --------------------------------------------------
    # Fail test on HTTP error responses (4xx / 5xx)
    # --------------------------------------------------
    if response.status_code >= 400:
        error_message = None

        try:
            error_body = response.json()
            error_message = error_body.get("errors") or error_body.get("message")
        except ValueError:
            error_body = response.text

        pytest.fail(
            f"API request failed.\n"
            "Status Code: {response.status_code}\n"
            f"Response: {error_body}"
        )
    
    check_count = test_data.get("check_count")

    info = None
    if method in ("GET", "POST", "PUT", "PATCH"):
        try:
            info = response.json()
        except ValueError:
            if is_valid_check(check_count):
                pytest.fail("Response is not valid JSON")

    if is_valid_check(check_count) and info is not None:
        # Find the first list in the response
        entries = next(
            (value for value in info.values() if isinstance(value, list)),[]
        )
        count = len(entries)
        allure.attach(str(count), "Total Entries Returned", allure.attachment_type.TEXT)
        if is_valid_check(check_count):
            assert count >= int(check_count)

    # Any number of check_elements can be added into the Excel spreadhseet.Make sure the heading starts with check_element
    for key, raw_value in test_data.items():
        if key.startswith("check_element"):
            value = str(raw_value).strip('"')

            if is_valid_check(value) and info is not None:
                with allure.step(f"Checking presence of '{value}' in response"):
                    assert value in json.dumps(info), f"`{value}` not found in response"


    allure.dynamic.title(f"{test_case_id} - {method} {base_url}")
    allure.dynamic.description("Unified authentication API validation")
